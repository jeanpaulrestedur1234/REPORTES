import tkinter as tk
import csv
from openpyxl import load_workbook
from CODIGOS.ejecutar import abrirArchivo

def guardar_datos(libro_excel,datos,hoja):
    hoja = libro_excel[hoja]
    for indice_fila, fila in enumerate(datos, start=1):
         for indice_columna, valor in enumerate(fila, start=1):
               celda = hoja.cell(row=indice_fila, column=indice_columna)  # Segunda columna en adelante
               celda.value = valor
def crear_excel():
    information=[]
    tabladata=tabla.get()

    inftabla = list(csv.reader(tabladata.strip().split('\n'), delimiter='\t'))
    inftabla=[[element for element in row if element!=''] for row in inftabla]
    
    for row in inftabla:
        rows=[]
        for element in row:
            try:
                rows.append(float(element))
                print(element)
            except:
                rows.append(element)
        
        information.append(rows)


    libro_excel = load_workbook('formato6min.xlsx')
    recorridos=[]
    hoja_prin=libro_excel['Datos generales']
    h=0
    for entry in entry_list:
        valor = entry.get()
        hoja_prin['B'+str(15+h)]=int(valor)
        h=h+1

    guardar_datos(libro_excel,datos,'Datos por tramo')
    guardar_datos(libro_excel,information,'Datos generales')
    folder=abrirArchivo()

    libro_excel.save(folder+'/Data_6Minutes.xlsx')
    ventana.destroy()
    # Aquí puedes agregar el código para crear el archivo de Excel
def cargar_info():
    entrada1=recorridoinfo.get()
    informacion = list(csv.reader(entrada1.strip().split('\n'), delimiter='\t'))

    informacion[7:-2]=[[float(element) for element in row if element!=''] for row in informacion[7:-2]]
   

    datos.extend(informacion[:-2])
    n_recorrido[0]+=1
    numrecorrido.config(text="Recorrido numero " +str(n_recorrido[0]))

    recorridoinfo.delete(0, tk.END)
 

def crear_ventana6min():
    global datos, numrecorrido, n_recorrido,recorridoinfo, entry_list,tabla,ventana
    n_recorrido = [1]
    ventana = tk.Tk()
    ventana.title("Crear Excel 6 min")
    ventana.configure(bg="light blue")

    # Crear el título
    titulo = tk.Label(ventana, text="Crear Excel 6 min", bg="light blue", font=("Arial", 16))
    titulo.grid(row=0, column=0, columnspan=2, pady=10)

    # Crear los entry y los textos
    entry_list = []
    for i in range(1, 7):
        entry = tk.Entry(ventana)
        entry.grid(row=i, column=0, padx=10, pady=5)
        texto = tk.Label(ventana, text="Minuto " + str(i), bg="light blue")
        texto.grid(row=i, column=1, padx=10, pady=5, sticky="W")
        entry_list.append(entry)

    # Crear los entry y los textos en otra columna
    tabla = tk.Entry(ventana)
    tabla.grid(row=1, column=2, padx=10, pady=5)
    texto = tk.Label(ventana, text="Datos de la tabla", bg="light blue")
    texto.grid(row=1, column=3, padx=10, pady=5, sticky="W")
    datos = []
    recorridoinfo = tk.Entry(ventana)
    recorridoinfo.grid(row=2, column=2, padx=10, pady=5)
    texto = tk.Label(ventana, text="Info del recorrido", bg="light blue")
    texto.grid(row=2, column=3, padx=10, pady=5, sticky="W")
    numrecorrido = tk.Label(ventana, text="Recorrido numero 1", bg="light blue")
    numrecorrido.grid(row=5, column=2, columnspan=2, pady=10)

    boton_cargar = tk.Button(ventana, text="Cargar", command=cargar_info)
    boton_cargar.grid(row=4, column=2, columnspan=2, pady=10)

    # Botón para crear el Excel
    boton_crear_excel = tk.Button(ventana, text="Crear Excel", command=crear_excel)
    boton_crear_excel.grid(row=8, column=0, columnspan=2, pady=10)

    # Iniciar el bucle de eventos
    ventana.mainloop()
