import tkinter as tk
import pandas as pd
import csv
from openpyxl import load_workbook
import numpy as np
from CODIGOS.ejecutar import abrirArchivo

def guardar_datos(libro_excel,datos,hoja):
    hoja = libro_excel[hoja]
    for indice_fila, fila in enumerate(datos, start=3):
         for indice_columna, valor in enumerate(fila, start=1):
               celda = hoja.cell(row=indice_fila, column=indice_columna)  # Segunda columna en adelante
               celda.value = valor

def guardar_valores():
    libro_excel = load_workbook('formato.xlsx')
    OA1=float(cuadro1.get())
    OA2=float(cuadro2.get())    
    entrada1=cuadro5.get()
    entrada2=cuadro6.get()    
    datos1 = list(csv.reader(entrada1.strip().split('\n'), delimiter='\t'))
    datos2 = list(csv.reader(entrada2.strip().split('\n'), delimiter='\t'))    
    dataOA1=[[float(element) for element in row  if element!='' ] for row in datos1[2:]]
    dataOA2=[[float(element) for element in row if element!=''] for row in datos2[2:]]
    guardar_datos(libro_excel,dataOA1,'OA1')
    guardar_datos(libro_excel,dataOA2,'OA2')
    GLO=libro_excel['Glo']
    GLO['B2']=OA1
    GLO['B3']=OA2
    GLO['B4']=(OA1+OA2)/2

    
    try:
         OC1=float(cuadro3.get())
         OC2=float(cuadro4.get())
         entrada3=cuadro7.get()
         entrada4=cuadro8.get()
         datos3 = list(csv.reader(entrada3.strip().split('\n'), delimiter='\t'))
         datos4 = list(csv.reader(entrada4.strip().split('\n'), delimiter='\t'))
         dataOC1=[[float(element) for element in row if element!=''] for row in datos3[2:]]
         dataOC2=[[float(element) for element in row if element!=''] for row in datos4[2:]]
         guardar_datos(libro_excel,dataOC1,'OC1')
         guardar_datos(libro_excel,dataOC2,'OC2')
         GLO['B5']=OC1
         GLO['B6']=OC2
         GLO['B7']=(OC1+OC2)/2

    except:
        hoja = libro_excel['OC1']
        hoja.delete_rows(1, hoja.max_row)
        hoja = libro_excel['OC2']
        hoja.delete_rows(1, hoja.max_row)

        pass

    directorio=abrirArchivo()

    libro_excel.save(directorio+'/stabilo.xlsx')
    ventana.destroy()



def guardar_stabilo():
    global cuadro1, cuadro2, cuadro3, cuadro4, cuadro5, cuadro6, cuadro7, cuadro8,ventana
    ventana = tk.Tk()
    ventana.title('CREAR EXCEL DE LA ESTABILOMETRIA')
    ventana.configure(bg="light blue")

    # Crear el título centrado
    titulo = tk.Label(ventana, text="Crear Excel de Estabilometría", bg="light blue", font=("Arial", 16))
    titulo.grid(row=0, column=0, columnspan=4, pady=10)

    # Crear los cuadros editables
    cuadro1 = tk.Entry(ventana)
    cuadro2 = tk.Entry(ventana)
    cuadro3 = tk.Entry(ventana)
    cuadro4 = tk.Entry(ventana)
    cuadro5 = tk.Entry(ventana)
    cuadro6 = tk.Entry(ventana)
    cuadro7 = tk.Entry(ventana)
    cuadro8 = tk.Entry(ventana)

    # Crear las etiquetas de texto
    texto1 = tk.Label(ventana, text="Ojos abiertos 1", bg="light blue")
    texto2 = tk.Label(ventana, text="Ojos abiertos 2", bg="light blue")
    texto3 = tk.Label(ventana, text="Ojos cerrados 1", bg="light blue")
    texto4 = tk.Label(ventana, text="Ojos cerrados 2", bg="light blue")
    texto5 = tk.Label(ventana, text="Datos:", bg="light blue")
    texto6 = tk.Label(ventana, text="Datos:", bg="light blue")
    texto7 = tk.Label(ventana, text="Datos:", bg="light blue")
    texto8 = tk.Label(ventana, text="Datos:", bg="light blue")

    # Crear el botón
    boton_guardar = tk.Button(ventana, text="crear excel", command=guardar_valores)

    # Ubicar los elementos en la ventana utilizando grid
    titulo.grid(row=0, column=0, columnspan=4, pady=10, padx=50)
    texto1.grid(row=1, column=0, sticky="E")
    cuadro1.grid(row=1, column=1, padx=10, pady=5)
    texto2.grid(row=2, column=0, sticky="E")
    cuadro2.grid(row=2, column=1, padx=10, pady=5)
    texto3.grid(row=3, column=0, sticky="E")
    cuadro3.grid(row=3, column=1, padx=10, pady=5)
    texto4.grid(row=4, column=0, sticky="E")
    cuadro4.grid(row=4, column=1, padx=10, pady=5)
    texto5.grid(row=1, column=2, sticky="E")
    cuadro5.grid(row=1, column=3, padx=10, pady=5)
    texto6.grid(row=2, column=2, sticky="E")
    cuadro6.grid(row=2, column=3, padx=10, pady=5)
    texto7.grid(row=3, column=2, sticky="E")
    cuadro7.grid(row=3, column=3, padx=10, pady=5)
    texto8.grid(row=4, column=2, sticky="E")
    cuadro8.grid(row=4, column=3, padx=10, pady=5)



    ventana.columnconfigure(0, weight=1)
    ventana.columnconfigure(1, weight=1)
    ventana.columnconfigure(2, weight=1)
    ventana.columnconfigure(3, weight=1)
    ventana.rowconfigure(0, weight=1)
    ventana.rowconfigure(1, weight=1)
    ventana.rowconfigure(2, weight=1)
    ventana.rowconfigure(3, weight=1)
    ventana.rowconfigure(4, weight=1)
    ventana.rowconfigure(5, weight=1)
    boton_guardar.grid(row=6, column=0, columnspan=4,padx=10, pady=5)

    # Iniciar el bucle principal de la ventana
    ventana.mainloop()


