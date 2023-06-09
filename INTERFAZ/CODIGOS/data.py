import pandas as pd
import os
from CODIGOS.extractor import*
import yaml
from yaml import SafeLoader
import openpyxl
import pandas as pd
from CODIGOS.similarity import cmc
from tkinter import filedialog
from tkinter import*
import tkinter as tk
import openpyxl

from tkinter import ttk
import numpy as np



def prepare_dataes(df):
    df = df.drop(0)
    df.columns = ['Bari Body X', 'Bari Body Y', 'Bari Left X', 'Bari Left Y', 'Bari Right X', 'Bari Right Y']

    return df


def read_data(comp=False, folder=str): 
    if not comp:
        file_data =''.join([folder,'/stabilo.xlsx'])
    else:
        file_data =filedialog.askopenfilename(title='cargar el excel de la estabilometria anterior')

    OA1 = pd.read_excel(file_data, sheet_name="OA1")
    OA2 = pd.read_excel(file_data, sheet_name="OA2")
    OC1 = pd.read_excel(file_data, sheet_name="OC1")
    OC2 = pd.read_excel(file_data, sheet_name="OC2")
    GLO = pd.read_excel(file_data, sheet_name="Glo")
    # print(OA1.drop(0), OA2.drop(0), prepare_data(OA1), prepare_data(OA2))
    if OC1.empty | OC2.empty:
        return prepare_dataes(OA1), prepare_dataes(OA2), pd.DataFrame(), pd.DataFrame(),  GLO
    return prepare_dataes(OA1), prepare_dataes(OA2), prepare_dataes(OC1), prepare_dataes(OC2), GLO


def prepare_data(df):
    df.dropna(axis=0, thresh=3, inplace=True)
    df.columns = df.iloc[0,:]
    df = df[df != df.columns].dropna(axis=0, thresh = 5).reset_index(drop=True)
    return df


def read_data6min(folder=str):

    general_data = pd.read_excel(folder + '/' +'Data_6Minutes.xlsx', sheet_name='Datos generales').dropna(thresh=2).drop(5).reset_index(drop=True)
    section_data = pd.read_excel(folder + '/' +'Data_6Minutes.xlsx', sheet_name='Datos por tramo')
    
    # Saving sec_data in new sheet
    
    path = os.getcwd()   
    writer = pd.ExcelWriter(folder + '/' + 'Data_6Minutes.xlsx', engine = 'openpyxl', mode = 'a',  if_sheet_exists = 'replace')
    prepare_data(section_data).to_excel(writer, sheet_name = 'Datos Completos')
    writer.save()
    return general_data, prepare_data(section_data)


def dh_study(s: str, study:str, path):
     if study=='Ml':
         study='M'

     if s =='DH_Children':
        sheet_name = f'DH{study}_Children'
     else:
        sheet_name = f'DH{study}_Adults'
     dh_data = pd.read_excel(os.sep.join(['C:/Users/marcha/Desktop/INTERFAZ/KINEMATICINFO/normatives','dh_normal.xlsx']), sheet_name = sheet_name,  header = 7, index_col = 0).dropna()

     return dh_data


def read_dataRMS(age: int, p_path: str, run_selected: int) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    global elect
    '''read_data This Function 

    Parameters
    ----------
    age :
        patient age
    p_path : 
        patient path
    run_selected : 
        the run selected for individual report
        
    Returns
    -------
    tuple(pd.Dataframe*6)
        Returns all dataframes from the txt files.
    '''
    columns = [
        'Frame', 'Time', 'Recto Femoral Derecho', 'Semitendinoso Derecho', 'Tibial anterior Derecho', 'Gastronemio Derecho', 
        'Recto Femoral Izquierdo', 'Semitendinoso Izquierdo', 'Tibial anterior Izquierdo', 'Gastronemio Izquierdo'
    ]
    print(f'{p_path}{os.sep}{run_selected}')

    try:
        emg_data=extract_emg(f'{p_path}{os.sep}{run_selected}')
    except:
        emg_data=None
    static=extract_static(f'{p_path}{os.sep}{run_selected}')
    times=static['time']
    scalars=static['scalar']
    events=static['event']
    

  
    kin_data=extract_angles(f'{p_path}{os.sep}{run_selected}')

    try:
        forces = extract_info(f'{p_path}{os.sep}{run_selected}')['force1d']
        momentos = extract_info(f'{p_path}{os.sep}{run_selected}')['torque1d']
        powers = extract_info(f'{p_path}{os.sep}{run_selected}')['power']

        thereare_kinetics = True
    except:
        forces=0
        momentos=0
        powers=0
        thereare_kinetics = False



    if age <= 16:
        sheet_name = 'DH_Children'

    else:
        sheet_name = 'DH_Adults'
    dh = pd.read_excel(os.sep.join(['normatives','dh_normal.xlsx']), sheet_name = sheet_name,  header = 7, index_col = 0).dropna()
    sheet_name = 'DH_Children'


    
    return times, scalars, events, kin_data, emg_data, dh,forces,powers,momentos,thereare_kinetics,sheet_name


async def save_excel(age,six, selected_run,folder,file_name): 
    p_path = folder
    name=' '.join(file_name.split()[:-1])
    s_name=' '.join(name.split()[1:])

   
    runs_mdx=[[i  for i in os.listdir(p_path) if 'Walking'in i]][0]
  


    for i in runs_mdx:
        if f'{str(selected_run)}.mdx' in i:
            select=i
    


    with open(os.sep.join(['config.yaml']), encoding='utf8') as yaml_file:
        config = yaml.load(yaml_file, Loader=SafeLoader)

        # Headers for excel file 
        
        kinematics_headers = config['Headers']['Kinematics']
        forces_headers = config['Headers']['Kinetics']['Force']
        torques_headers = config['Headers']['Kinetics']['Torque']
        power_headers = config['Headers']['Kinetics']['Power']

        # Filenames of data to read 

        filenames = config['Filenames']

        # Path where the six min data was saved
    
        six_data_path = config['sixmin_folder_path']
    

    
    
    try:
        forces = extract_info(f'{p_path}{os.sep}{select}')['force1d']
        torques = extract_info(f'{p_path}{os.sep}{select}')['torque1d']
        powers = extract_info(f'{p_path}{os.sep}{select}')['power']            
        thereare_kinetics = True

    except :

        thereare_kinetics = False

    data_similarity = cmc(age,p_path,runs_mdx)
    df2save=data_similarity
  



    
    analisis=[np.count_nonzero(df2save > 0.75)/np.size(df2save) *100,np.count_nonzero((0.75>=df2save) & (df2save>0.5))/np.size(df2save) *100,np.count_nonzero((0.5>=df2save) & (df2save>0.25))/np.size(df2save) *100,
                        np.count_nonzero(0.25>=df2save )/np.size(df2save) *100]


    if six.lower() == 'si':


        # six_data_path = r'C:\\Users\\marcha\\Desktop\\PACIENTES\\PACIENTES VAR_6_MINS'
        p_folder_name = ' '.join(file_name.split()[:-1]) + ' 6_MIN'
        p_six_path = os.sep.join([six_data_path, p_folder_name, f'Data6_min_{s_name}.xlsx'])
        print(p_six_path)



        # Open six_min_data excel

        try: 

            excel_book = openpyxl.load_workbook(p_six_path)

        except FileNotFoundError:
            try: 
                p_folder_name = name + ' 6MIN'
                p_six_path = os.sep.join([six_data_path, p_folder_name, f'Data6_min_{s_name}.xlsx'])   
                excel_book = openpyxl.load_workbook(p_six_path)
            except FileNotFoundError:
                print('NO SE ENCONTRO EL LIBRO')
                excel_book = Workbook()


        # Create new sheets


        excel_book = add_new_sheet(excel_book, data_similarity, 'Similitud') 
        excel_filename = f'{p_path}{os.sep}Grafica&Data_6min_{s_name}.xlsx'

    elif six.lower() == 'no':


        # Create new workbook
        
        excel_book = Workbook()
        excel_book = add_new_sheet(excel_book, data_similarity, 'Similitud',addnew=False)
        excel_filename = f'{p_path}{os.sep}Graficas_{s_name}.xlsx'
    else:
        ventana_secundaria = tk.Toplevel()
        ventana_secundaria.title("advertencia")
        ventana_secundaria.config(width=300, height=200)
    # Crear un botón dentro de la ventana secundaria
    # para cerrar la misma.
        boton_cerrar = ttk.Button(ventana_secundaria,text="Cerrar ventana", command=ventana_secundaria.destroy)
        boton_cerrar.place(x=75, y=75)
        adv=ttk.Label(ventana_secundaria,text='Asegurese de haber marcado bien los 6m')
        adv.place(x=50,y=50)

    kin_data = extract_angles(f'{p_path}{os.sep}{select}')
    kin_data.columns=[kinematics_headers[joint] for joint in kin_data.columns]

    
    excel_book=add_new_sheet(excel_book,kin_data,'Datos Cinematica')
    
    if thereare_kinetics:
        forces.columns = [forces_headers[joint] for joint in forces.columns]
        torques.columns = [torques_headers[joint] for joint in torques.columns]
        powers.columns = [power_headers[joint] for joint in powers.columns]
        
        # Create new sheets

        excel_book = add_new_sheet(excel_book, forces, 'Fuerzas')
        excel_book = add_new_sheet(excel_book, torques, 'Torques')
        excel_book = add_new_sheet(excel_book, powers, 'Potencias')

    
    sheet =excel_book['Similitud']
    evaluacion=['Excelente','Bueno','Deficiente','Malo']
    sheet['B26'].alignment =  Alignment(horizontal='center', wrap_text=True)
    sheet['B26']='Porcentaje de los datos'
    for i in range(4):
         sheet[''.join(['A',str(27+i)])]=evaluacion[i]
         sheet[''.join(['B',str(27+i)])]=''.join([str(round(analisis[i],2)),' %'])
         sheet[''.join(['B',str(27+i)])].alignment =  Alignment(horizontal='center', wrap_text=True)
    try:
        statics=extract_static(f'{p_path}{os.sep}{select}')['track1d']
        maxswing=[statics['MaxRswing'][0],statics['MaxLswing'][0]]
        sheet['D27']="Altura maxima izquierdo (cm)"
        sheet['D28']="Altura maxima Derecho (cm)"
        sheet['E27']=maxswing[1]*100
        sheet['E28']=maxswing[0]*100
    except:
        pass
 


    excel_book.save(excel_filename)
  
 

    return select