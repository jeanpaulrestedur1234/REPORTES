import xml.etree.ElementTree as ET
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

def add_new_sheet(excel_book, df, sheetname, addnew = True):
    if addnew:
        excel_book.create_sheet(sheetname)
        ws = excel_book[sheetname]
    else: 
        ws = excel_book.active
        ws.title = sheetname

    rows = dataframe_to_rows(df)

    # For each row

    for r_idx, row in enumerate(rows, 1):

        # Write each cell for each column
        
        
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for col in ws.columns:
        column = col[0].column_letter
        ws.column_dimensions[column].width = 20 if sheetname != 'Similitud' else 35
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment =  Alignment(horizontal='center', wrap_text=True)
    return excel_book


def extract_info(name):
  tree = ET.parse(name)
  root = tree.getroot()
  element=[]
  for elem in root:

    for subelem in elem:
        element.append(subelem.tag)
  buscador=dict(zip(element,range(len(element))))

  cycle_info=[]
  for elem in root[0][buscador['cycle']]:
     if 'band' in elem.tag:
       pass
     else:
          cycle_info.append(elem.tag)
  cycle_info=np.unique(cycle_info)


  cycle_data={}
  for i in cycle_info:
    elemento={}
    for elem in root[0][buscador['cycle']]:
      if  elem.tag==i:
        if '1E+09'==elem.attrib['scaleFactor']:
          scale=1000000
        else:
          scale=int(elem.attrib['scaleFactor'])
        data=[]
        if elem.attrib['data']=='':
           inicio=np.NAN
        
        
        else:
           inicio=int(elem.attrib['data'].split()[2])
    
        
     


        if inicio==1:
            data.append(np.NaN)

        for k in elem.attrib['data'].split()[3:]:

          
          try:
            data.append(int(k)/scale)
          except:
            data.append(np.NAN)
        
        if len(data)<101:
           data=data+[np.NAN]*(101-len(data))


 
        nombre=elem.attrib['label']

        elemento[nombre]=data
      
  
    cycle_data[i]=pd.DataFrame.from_dict(elemento, orient='index').transpose()
  
  return cycle_data



def extract_angles(name):
    tree = ET.parse(name)
    root = tree.getroot()

    element = [subelem.tag for elem in root for subelem in elem]
    buscador = {element[i]: i for i in range(len(element))}

    elemento = {}

    for elem in root[0][buscador['cycle']]:
        if elem.tag == 'angle':
            scaleFactor = elem.attrib['scaleFactor']
            scale = 10**9 if scaleFactor == '1E+09' else int(scaleFactor)

            data = []
            for k in elem.attrib['data'].split()[3:]:
                try:
                    data.append(int(k) / scale)
                except:
                    data.append(np.NAN)

            nombre = elem.attrib['label']
            elemento[nombre] = data

    angulos = pd.DataFrame.from_dict(elemento, orient='index').transpose()

    return angulos


def extract_emg(name):
    tree = ET.parse(name)
    root = tree.getroot()

    element = [subelem.tag for elem in root for subelem in elem]
    ubicaciones = [i for i, tag in enumerate(element) if tag == 'stream']

    ubicacion = None
    emg = None

    for i in ubicaciones:
        stream_info = [elem.tag for elem in root[0][i]]
        stream_info = np.unique(stream_info)

        if 'emg' in stream_info:
            ubicacion = i
            emg = stream_info
            break

    stream = {}
    for i in emg:
        elemento = {}
        for elem in root[0][ubicacion]:
            if elem.tag == i:
                scaleFactor = elem.attrib['scaleFactor']
                scale = 10**6 if scaleFactor == '1E+09' else int(scaleFactor)

                data = []
                for k in elem.attrib['data'].split()[2:]:
                    try:
                        data.append(int(k) / scale)
                    except:
                        data.append(np.NAN)

                nombre = elem.attrib['label']
                elemento[nombre] = data

        stream[i] = pd.DataFrame.from_dict(elemento, orient='index').transpose()

    return stream['emg']


def extract_static(name):
    tree = ET.parse(name)
    root = tree.getroot()

    element = [subelem.tag for elem in root for subelem in elem]
    buscador = {element[i]: i for i in range(len(element))}

    static_info = [elem.tag for elem in root[0][buscador['static']]]
    static_info = np.unique(static_info)

    static = {}
    ah = {}
    for i in static_info:
        elemento = {}
        for elem in root[0][buscador['static']]:
            if elem.tag == i:
                scaleFactor = elem.attrib['scaleFactor']
                scale = 1000000 if scaleFactor == '1E+09' else int(scaleFactor)

                try:
                    data = [int(elem.attrib['data']) / scale]
                except:
                    data = [int(i) / scale for i in elem.attrib['data'].replace('I', '').split()]

                nombre = elem.attrib['label']
                elemento[nombre] = data

        static[i] = pd.DataFrame.from_dict(elemento, orient='index').transpose()
        ah[i] = elemento

    return static

